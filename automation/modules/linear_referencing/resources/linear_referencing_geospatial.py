import csv
from io import StringIO, BytesIO
from flask import request
from flask_restx import Namespace, Resource, fields, reqparse
from werkzeug.datastructures import FileStorage
from .... import PyAutomation
from ....extensions.api import api
from ....extensions import _api as Api


ns = Namespace('Linear Referencing Geospatial', description='Linear referencing geospatial CRUD and interpolation')
app = PyAutomation()


create_point_model = api.model("create_linear_referencing_geospatial_model", {
    'segment_name': fields.String(required=True, description='Pipeline segment name'),
    'kp': fields.Float(required=True, description='Kilometer Post'),
    'latitude': fields.Float(required=True, description='Latitude (WGS84, Y)'),
    'longitude': fields.Float(required=True, description='Longitude (WGS84, X)'),
    'elevation': fields.Float(required=False, description='Elevation (meters above sea level)')
})

update_point_model = api.model("update_linear_referencing_geospatial_model", {
    'segment_name': fields.String(required=False, description='Pipeline segment name'),
    'kp': fields.Float(required=False, description='Kilometer Post'),
    'latitude': fields.Float(required=False, description='Latitude (WGS84, Y)'),
    'longitude': fields.Float(required=False, description='Longitude (WGS84, X)'),
    'elevation': fields.Float(required=False, description='Elevation (meters above sea level)')
})

interpolate_model = api.model("interpolate_linear_referencing_geospatial_model", {
    'segment_name': fields.String(required=True, description='Pipeline segment name'),
    'kp': fields.Float(required=True, description='Requested KP')
})


@ns.route('/')
class LinearReferencingCollection(Resource):
    parser = reqparse.RequestParser()
    parser.add_argument('segment_name', type=str, location='args', help='Optional segment name filter')

    @api.doc(security='apikey', description="Retrieves all points or points by segment.")
    @api.response(200, "Success")
    @Api.token_required(auth=True)
    @ns.expect(parser)
    def get(self):
        args = self.parser.parse_args()
        segment_name = args.get("segment_name")
        if segment_name:
            data = app.get_linear_referencing_geospatial_points_by_segment(segment_name=segment_name)
            return {"data": data}, 200
        return {"data": app.get_linear_referencing_geospatial_points()}, 200


@ns.route('/<int:point_id>')
@api.param('point_id', 'Linear referencing geospatial point ID')
class LinearReferencingPointResource(Resource):

    @api.doc(security='apikey', description="Retrieves one geospatial point by ID.")
    @api.response(200, "Success")
    @api.response(404, "Point not found")
    @Api.token_required(auth=True)
    def get(self, point_id:int):
        point = app.get_linear_referencing_geospatial_point(id=point_id)
        if point is None:
            return {"message": f"Linear referencing geospatial point {point_id} not found"}, 404
        return {"data": point}, 200

    @api.doc(security='apikey', description="Updates one geospatial point by ID.")
    @api.response(200, "Updated")
    @api.response(400, "Invalid payload")
    @api.response(404, "Point not found")
    @Api.token_required(auth=True)
    @ns.expect(update_point_model)
    def put(self, point_id:int):
        payload = api.payload or {}
        if not payload:
            return {"message": "No fields to update provided"}, 400
        data, message = app.update_linear_referencing_geospatial_point(id=point_id, **payload)
        if data is None:
            if "not found" in message.lower():
                return {"message": message}, 404
            return {"message": message}, 400
        return {"message": message, "data": data}, 200

    @api.doc(security='apikey', description="Deletes one geospatial point by ID.")
    @api.response(200, "Deleted")
    @api.response(404, "Point not found")
    @Api.token_required(auth=True)
    def delete(self, point_id:int):
        success, message = app.delete_linear_referencing_geospatial_point(id=point_id)
        if not success:
            if "not found" in message.lower():
                return {"message": message}, 404
            return {"message": message}, 400
        return {"message": message}, 200


@ns.route('/add')
class LinearReferencingCreateResource(Resource):

    @api.doc(security='apikey', description="Creates a new geospatial linear-referencing point.")
    @api.response(200, "Created")
    @api.response(400, "Creation error")
    @Api.token_required(auth=True)
    @ns.expect(create_point_model)
    def post(self):
        payload = api.payload or {}
        point, message = app.create_linear_referencing_geospatial(
            segment_name=payload.get("segment_name"),
            kp=payload.get("kp"),
            latitude=payload.get("latitude"),
            longitude=payload.get("longitude"),
            elevation=payload.get("elevation")
        )
        if point is None:
            return {"message": message}, 400
        return {"message": message, "data": point}, 200


@ns.route('/interpolate')
class LinearReferencingInterpolateResource(Resource):

    @api.doc(security='apikey', description="Retrieves interpolated geospatial coordinates by segment and KP.")
    @api.response(200, "Success")
    @api.response(400, "Interpolation error")
    @Api.token_required(auth=True)
    @ns.expect(interpolate_model)
    def post(self):
        payload = api.payload or {}
        data, message = app.get_geospatial_by_segment_and_kp(
            segment_name=payload.get("segment_name"),
            kp=payload.get("kp")
        )
        if data is None:
            return {"message": message}, 400
        return {"message": message, "data": data}, 200


bulk_import_parser = reqparse.RequestParser()
bulk_import_parser.add_argument(
    'file',
    type=FileStorage,
    location='files',
    required=True,
    help='CSV or XLSX file to import'
)
bulk_import_parser.add_argument(
    'segment_name',
    type=str,
    location='form',
    required=False,
    help='Default segment name (used when the file does not include a segment_name column)'
)
bulk_import_parser.add_argument(
    'update_existing',
    type=str,
    location='form',
    required=False,
    default='true',
    help='Update existing points if they already exist (true/false, default: true)'
)


@ns.route('/bulk_import')
class LinearReferencingBulkImportResource(Resource):

    @api.doc(
        security='apikey',
        description="Imports a full profile from CSV/XLSX file. "
                    "Expected columns: segment_name (optional if sent in form), kp, latitude, longitude, elevation (optional)."
    )
    @api.response(200, "Import processed")
    @api.response(400, "Invalid file or payload")
    @Api.token_required(auth=True)
    @ns.expect(bulk_import_parser)
    def post(self):
        if "file" not in request.files:
            return {"message": "file is required (CSV or XLSX)"}, 400

        upload = request.files["file"]
        filename = (upload.filename or "").strip()
        if not filename:
            return {"message": "Invalid filename"}, 400

        default_segment_name = request.form.get("segment_name")
        update_existing_raw = (request.form.get("update_existing", "true") or "true").strip().lower()
        update_existing = update_existing_raw in ("1", "true", "yes", "y")

        try:
            rows = self._parse_rows(upload=upload, filename=filename)
        except Exception as err:
            return {"message": f"Invalid file format: {str(err)}"}, 400

        result = app.import_linear_referencing_profile(
            rows=rows,
            default_segment_name=default_segment_name,
            update_existing=update_existing
        )
        return {"data": result}, 200

    def _normalize_header(self, value:str)->str:
        return (value or "").strip().lower()

    def _canonicalize_row(self, row:dict)->dict:
        mapping = {}
        for key, value in row.items():
            normalized = self._normalize_header(key)
            mapping[normalized] = value

        segment_name = (
            mapping.get("segment_name")
            or mapping.get("segment")
            or mapping.get("segmento")
        )
        kp = mapping.get("kp")
        latitude = (
            mapping.get("latitude")
            or mapping.get("lat")
            or mapping.get("y")
        )
        longitude = (
            mapping.get("longitude")
            or mapping.get("lon")
            or mapping.get("lng")
            or mapping.get("x")
        )
        elevation = (
            mapping.get("elevation")
            or mapping.get("elev")
            or mapping.get("altitude")
            or mapping.get("altura")
        )

        return {
            "segment_name": segment_name,
            "kp": kp,
            "latitude": latitude,
            "longitude": longitude,
            "elevation": elevation
        }

    def _parse_rows(self, upload, filename:str)->list[dict]:
        lowered = filename.lower()
        if lowered.endswith(".csv"):
            raw_text = upload.stream.read().decode("utf-8-sig")
            reader = csv.DictReader(StringIO(raw_text))
            return [self._canonicalize_row(row) for row in reader]

        if lowered.endswith(".xlsx"):
            try:
                from openpyxl import load_workbook
            except Exception as err:
                raise ValueError(f"openpyxl is required for XLSX import: {str(err)}")

            content = upload.stream.read()
            wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = next(rows_iter, None)
            if headers is None:
                return []
            header_names = [str(h) if h is not None else "" for h in headers]

            parsed = []
            for values in rows_iter:
                row_dict = {}
                for idx, header in enumerate(header_names):
                    row_dict[header] = values[idx] if idx < len(values) else None
                parsed.append(self._canonicalize_row(row_dict))
            return parsed

        raise ValueError("Only .csv and .xlsx files are supported")
